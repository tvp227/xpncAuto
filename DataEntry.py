import openpyxl
import os
import getpass
from datetime import datetime

username = getpass.getuser()

# Load the Excel file
wb = openpyxl.load_workbook(f'C:/Users/{username}/Documents/XPNCv2/expenses.xlsx')

# Select the first sheet
sheet = wb.active

# Set the starting row for entries
start_row = 5

# Find the next empty row in column A
next_row = start_row

# Prompt for adding an entry
add_another = "yes"
while add_another.lower() == "yes":
    # Check if data already exists in a cell and notify before input
    existing_data = []
    for col in range(1, 7):
        cell_value = sheet.cell(row=next_row, column=col).value
        existing_data.append(cell_value if cell_value else "-")

    print(f"Existing data in row {next_row}:")
    print(f"Date: {existing_data[0]}")
    print(f"Description: {existing_data[1]}")
    print(f"Reason: {existing_data[2]}")
    print(f"Receipt: {existing_data[3]}")
    print(f"Cost: {existing_data[4]}")
    print(f"Quantity: {existing_data[5]}")

    # Prompt for inputs
    date = input("Enter the date (leave blank to keep it): ") or existing_data[0]
    description = input("Enter the description (leave blank to keep it): ") or existing_data[1]
    reason = input("Enter the reason (leave blank to keep it): ") or existing_data[2]
    receipt = input("Do you have a receipt? (leave blank to keep it): ") or existing_data[3]
    cost = input("Enter the cost (leave blank to keep it): ") or existing_data[4]
    quantity = input("Enter the quantity (leave blank to keep it): ") or existing_data[5]

    # Fill in the values in separate cells across rows in column A
    sheet.cell(row=next_row, column=1).value = date
    sheet.cell(row=next_row, column=2).value = description
    sheet.cell(row=next_row, column=3).value = reason
    sheet.cell(row=next_row, column=4).value = receipt
    sheet.cell(row=next_row, column=5).value = cost
    sheet.cell(row=next_row, column=6).value = quantity

    next_row += 1

    # Prompt for adding another entry
    add_another = input("Do you want to add another entry? (yes/no): ")

# Add the current date and time to cell I2
current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
sheet.cell(row=2, column=9).value = current_datetime

# Save the changes
wb.save(f'C:/Users/{username}/Documents/XPNCv2/expenses.xlsx')
