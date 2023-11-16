import pytesseract
from PIL import Image
import re
import os
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askdirectory

# Set the path to the Tesseract executable
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Select the directory containing the images
Tk().withdraw()
images_dir = askdirectory(title='Select Images Directory')

# Set the path to the Excel template
template_path = r'C:\Users\tomvp\Documents\XPNCv2\expenses.xlsx'

# Load the Excel template
workbook = openpyxl.load_workbook(template_path)
sheet = workbook.active

# Set initial row counter
row = 5

# Iterate over each file in the directory
for filename in os.listdir(images_dir):
    # Check if the file is an image
    if filename.endswith('.jpg') or filename.endswith('.png'):
        # Create the full path to the image file
        image_path = os.path.join(images_dir, filename)

        try:
            # Open the image using PIL
            image = Image.open(image_path)

            # Perform OCR on the image
            text = pytesseract.image_to_string(image)

            # Extract price using regular expression
            price_match = re.search(r'£(\d+(\.\d{1,2})?)', text)
            price = '£' + price_match.group(1) if price_match else None

            # Extract date using regular expression
            date_match = re.search(r'(\d{1,2}-[A-Z]{3}-\d{2})', text)
            date = date_match.group(1) if date_match else None

            # Extract title
            title_match = re.search(r'TICKET|(&\s)(?!.*zones 1-6)(.*)', text)
            title = title_match.group(0) if title_match else None

            # Display the extracted information in the terminal
            print("Image:", filename)
            print("Price:", price)
            print("Date:", date)
            print("Title:", title)
            print("------------------------")

            # Write the extracted information to the Excel sheet
            sheet.cell(row=row, column=1).value = date
            sheet.cell(row=row, column=5).value = price
            sheet.cell(row=row, column=2).value = title

            # Increment row counter
            row += 1

        except Exception as e:
            print("Error processing image:", filename)
            print("Exception:", e)
            print("------------------------")

# Save the updated workbook
workbook.save(template_path)

# Print acknowledgment to the end user
print("Extraction completed. The data has been saved to", template_path)
